"""Genera el Excel descargable: un resumen por concesionario (equivalente
al bloque del 'Dealer Dashboard' original, pero para todos los
concesionarios a la vez) más el detalle mensual calculado desde la BBDD.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import dashboard, metrics

FUENTE = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FUENTE, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FUENTE, bold=True, size=14)
BASE_FONT = Font(name=FUENTE)
PCT_FORMAT = "0.0%"


def _write_header(ws, row, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def _autosize(ws, n_cols, width=16):
    for j in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(j)].width = width


def build_workbook(resumen, dealers, marca: str, periodo: str, mes_referencia: str | None = None) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Dashboard concesionarios"
    ws["A1"] = f"Seguimiento UC Retail & Wholesale — {marca} — {periodo}" + (f" ({mes_referencia})" if mes_referencia else "")
    ws["A1"].font = TITLE_FONT

    headers = [
        "Código", "Concesionario", "Grupo propietario",
        "Objetivo Retail", "Realizado Retail", "% Cumplimiento",
        "BPS/MN", "% BPS/MN",
        "Remarketing", "% Remarketing",
        "Objetivo BEV", "BEV", "% BEV",
        "Wholesale UC", "Wholesale YUC",
        "Mercado <6 años", "% Penetración",
        "Mystery Shopping",
    ]
    header_row = 3
    _write_header(ws, header_row, headers)

    col_bmw = "vende_bmw" if marca == "BMW" else "vende_mini"
    fila = header_row + 1
    for _, d in dealers[dealers[col_bmw] == "Si"].sort_values("concesionario").iterrows():
        codigo = int(d["codigo_dealer"])
        k = dashboard.kpi_bloque(resumen, codigo, marca, periodo, mes_referencia)
        valores = [
            codigo, d["concesionario"], d["grupo_propietario"],
            k["objetivo_retail"], k["realizado_retail"], None,
            k["bps"], None,
            k["remarketing"], None,
            k["objetivo_bev"], k["bev"], None,
            k["wholesale_uc"], k["wholesale_yuc"],
            k["mercado_menos_6_anos"], None,
            k["mystery_shopping"],
        ]
        for j, v in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=j, value=v)
            c.font = BASE_FONT

        # % como fórmulas locales (recalculan solas si se edita el valor a mano)
        ws.cell(row=fila, column=6, value=f"=IFERROR(E{fila}/D{fila},\"\")").number_format = PCT_FORMAT
        ws.cell(row=fila, column=8, value=f"=IFERROR(G{fila}/E{fila},\"\")").number_format = PCT_FORMAT
        ws.cell(row=fila, column=10, value=f"=IFERROR(I{fila}/E{fila},\"\")").number_format = PCT_FORMAT
        ws.cell(row=fila, column=13, value=f"=IFERROR(L{fila}/E{fila},\"\")").number_format = PCT_FORMAT
        ws.cell(row=fila, column=17, value=f"=IFERROR(E{fila}/P{fila},\"\")").number_format = PCT_FORMAT
        fila += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize(ws, len(headers))

    # Hoja de detalle mensual (equivalente a las pestañas "feeder" del Excel original)
    ws2 = wb.create_sheet("Detalle mensual")
    detalle = resumen[resumen["marca"] == marca].merge(
        dealers[["codigo_dealer", "concesionario", "grupo_propietario"]], on="codigo_dealer", how="left"
    )
    det_headers = ["codigo_dealer", "concesionario", "grupo_propietario", "mes", "retail", "bps", "remarketing", "bev", "wholesale_uc", "wholesale_yuc"]
    _write_header(ws2, 1, det_headers)
    for i, (_, r) in enumerate(detalle[det_headers].sort_values(["concesionario", "mes"]).iterrows(), start=2):
        for j, col in enumerate(det_headers, start=1):
            ws2.cell(row=i, column=j, value=r[col]).font = BASE_FONT
    _autosize(ws2, len(det_headers))
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
