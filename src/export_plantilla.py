"""Rellena la plantilla original de Excel del usuario
(`data/plantilla_seguimiento.xlsx`, el `SEGUIMIENTO UC RETAIL &
WHOLESALE` que nos pasó) con los datos que la app ya calculó -en vez
de generar un Excel desde cero-, para conservar el diseño original:
logos, colores, orden y agrupación por Distrito (ya viene así en la
plantilla, no hay que rehacerla).

Cómo funciona: en el Excel original, las celdas de "Realizado" salen
de fórmulas que apuntan a OTROS archivos que llegan cada mes por
separado ([2]BPS, [2]BEV, [2]RETAIL ORIGEN RMKT, [2]COMPRAS UC/YUC,
[9]BMW/MINI...) -no los tenemos, así que esas celdas no se pueden
recalcular solas. Este módulo sólo SOBRESCRIBE esas celdas concretas
con el valor que la app ya calculó desde la BBDD subida. El resto de
fórmulas (SUMIF locales entre pestañas del propio archivo, %, la vista
por Grupo Propietario) se dejan tal cual: son locales, así que Excel
las recalcula solas al abrir el archivo -la plantilla ya trae activado
"recalcular todo al abrir" (`fullCalcOnLoad`), no hay que forzarlo.

Si para un mes no hay datos en el archivo de ventas subido (todavía no
ha llegado ese mes), la celda se deja en blanco -nunca se deja el dato
antiguo que traía la plantilla, para no confundir un mes sin subir con
un mes en cero.

Cada hoja tiene además una celda "mes de referencia" (p.ej. BZ4 en
"UC BMW 2026 BPS") de la que dependen las columnas de "Acumulado Mes"
/ "Acumulado Mes 2S" -son fórmulas locales (CHOOSE/SUMPRODUCT) que sólo
necesitan que esa celda tenga el mes correcto para sumar bien; se
sincroniza con el mes hasta el que se quiere el acumulado en la
descarga. La pestaña "Dealer Dashboard" tiene su propia celda de mes
(G3) más una de concesión (E5) y periodo (F3): sus celdas de
Objetivo/Realizado/BPS/Remarketing/BEV (G12:G19, I12:I19) son fórmulas
de matriz (INDEX+MATCH+CHOOSE) cuyo resultado en caché se pierde al
volver a guardar el archivo con openpyxl -por eso, además de
sincronizar el mes, se sobrescriben esas celdas concretas con el
mismo cálculo que ya hace `dashboard.kpi_bloque()` para el
concesionario/periodo/mes que tenga seleccionados esa hoja, así se ve
bien aunque el visor no recalcule fórmulas de matriz. El resto de la
hoja (%, semáforos, tablas de bonificación, Mystery Shopping) son
fórmulas normales que sí se recalculan solas a partir de esas celdas."""
import io
from pathlib import Path

import openpyxl
import pandas as pd

from . import config, dashboard, storage

# Celda "mes de referencia" de la que dependen las columnas Acumulado
# Mes / Acumulado Mes 2S de cada hoja (encontradas explorando las
# fórmulas de esas columnas -ver comentario de arriba).
_CELDA_MES_REFERENCIA = {
    "UC BMW 2026 BPS": "BZ4",
    "UC MINI 2026 MINI NEXT": "CL4",
    "BEV BMW 2026": "AP4",
    "BEV MINI 2026": "AP4",
    "WHOLESALE BMW 2026": "AD4",
    "WHOLESALE MINI 2026": "AD4",
    "PENETRACION MERCADO VO BMW": "AP4",
    "PENETRACION MCDO VO MINI": "AP4",
    "UC BMW 2026": "AD2",
    "UC MINI 2026": "AD2",
}

# Traduce el texto del selector de periodo de "Dealer Dashboard" (F3)
# al nombre de periodo que ya entiende metrics.meses_de_periodo().
_PERIODO_DASHBOARD = {
    "ACUMULADO MES": "ACUMULADO MES",
    "ACUMULADO MES 2º S": "ACUMULADO MES 2S",
    "ANUAL": "ANUAL",
    "SEMESTRE 1": "SEMESTRE 1",
    "SEMESTRE 2": "SEMESTRE 2",
    **{m: m for m in config.MESES},
}

PLANTILLA_PATH = Path(__file__).resolve().parent.parent / "data" / "plantilla_seguimiento.xlsx"

# (hoja BMW, hoja MINI, etiqueta de la subcolumna a sobrescribir en la
# cabecera de la hoja). fila_mes / fila_subheader varían según la hoja.
_HOJAS_BPS_REMARKETING = {
    "BMW": ("UC BMW 2026 BPS", "BPS", "RETAIL ORIGEN RMK"),
    "MINI": ("UC MINI 2026 MINI NEXT", "M-NEXT", "RETAIL ORIGEN RMK"),
}
_HOJAS_BEV = {"BMW": "BEV BMW 2026", "MINI": "BEV MINI 2026"}
_HOJAS_WHOLESALE = {"BMW": "WHOLESALE BMW 2026", "MINI": "WHOLESALE MINI 2026"}
_HOJAS_RETAIL_OBJ = {"BMW": "UC BMW 2026", "MINI": "UC MINI 2026"}
_HOJAS_MERCADO = {"BMW": "PENETRACION MERCADO VO BMW", "MINI": "PENETRACION MCDO VO MINI"}
_HOJA_MYS = "MYS 2026"


def plantilla_disponible() -> bool:
    return PLANTILLA_PATH.exists()


def _localizar_columnas(ws, fila_mes: int, fila_subheader: int) -> dict:
    """Explora la cabecera de bloques mensuales fusionados (el nombre
    del mes sólo aparece en la primera celda del bloque) y devuelve
    {(MES, ETIQUETA): columna}, con MES/ETIQUETA en mayúsculas."""
    meses_col = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=fila_mes, column=c).value
        if isinstance(v, str) and v.strip().upper() in config.MESES:
            meses_col[c] = v.strip().upper()
    cols_ordenadas = sorted(meses_col)
    resultado = {}
    for idx, c0 in enumerate(cols_ordenadas):
        c1 = cols_ordenadas[idx + 1] if idx + 1 < len(cols_ordenadas) else ws.max_column + 1
        mes = meses_col[c0]
        for c in range(c0, c1):
            etiqueta = ws.cell(row=fila_subheader, column=c).value
            if etiqueta:
                resultado[(mes, str(etiqueta).strip().upper())] = c
    return resultado


def _filas_por_dealer(ws, col_codigo: int = 2, fila_inicio: int = 5) -> dict:
    """{codigo_dealer: fila}. Las filas de subtotal ("TOTAL DISTRITO
    X"...) se descartan solas porque su "código" no es un entero."""
    filas = {}
    for r in range(fila_inicio, ws.max_row + 1):
        v = ws.cell(row=r, column=col_codigo).value
        if isinstance(v, int):
            filas[v] = r
    return filas


def _escribir_mes(ws, filas_dealer: dict, columnas: dict, etiqueta: str, valores: dict, escala: float = 1.0):
    """valores: {(codigo_dealer, mes): numero}, sólo con las claves que
    SÍ tienen dato. Escribe TODAS las filas de dealer x los 12 meses
    -si a una combinación le falta valor, la celda se deja en blanco
    (None) en vez de dejar el número antiguo que traía la plantilla
    -Excel guarda el último valor calculado de una fórmula con vínculo
    externo aunque ese archivo ya no esté disponible, así que sin este
    blanqueo explícito se colaría un dato real de meses pasados-. Si
    esta plantilla no tiene columna para un mes (fuera de rango), esa
    combinación se salta sin tocar nada."""
    for codigo, fila in filas_dealer.items():
        for mes in config.MESES:
            col = columnas.get((mes, etiqueta))
            if col is None:
                continue
            valor = valores.get((codigo, mes))
            ws.cell(row=fila, column=col).value = None if valor is None else valor * escala


def _dict_resumen(resumen: pd.DataFrame, marca: str, columna: str) -> dict:
    """{(codigo_dealer, mes): valor} sólo para los meses presentes en
    `resumen` -el resto de combinaciones se blanquean en _escribir_mes."""
    sub = resumen[resumen["marca"] == marca]
    return {(int(r["codigo_dealer"]), r["mes"]): r[columna] for _, r in sub.iterrows()}


def _dict_manual(df: pd.DataFrame, marca: str, col_periodo: str) -> dict:
    """{(codigo_dealer, periodo): valor} con lo que haya guardado a
    mano (objetivo/mercado/mystery shopping) -el resto se blanquea en
    _escribir_mes, igual que _dict_resumen."""
    if df.empty:
        return {}
    sub = df[df["marca"] == marca]
    return {(int(r["codigo_dealer"]), r[col_periodo]): r["valor"] for _, r in sub.iterrows()}


def _sincronizar_mes_referencia(wb, mes_referencia: str) -> None:
    for hoja, celda in _CELDA_MES_REFERENCIA.items():
        if hoja in wb.sheetnames:
            wb[hoja][celda].value = mes_referencia


def _rellenar_dealer_dashboard(wb, resumen: pd.DataFrame, dealers: pd.DataFrame, remarketing_disponible: bool, mes_referencia: str) -> None:
    if "Dealer Dashboard" not in wb.sheetnames:
        return
    ws = wb["Dealer Dashboard"]
    ws["G3"].value = mes_referencia

    concesion = str(ws["E5"].value or "").strip().upper()
    fila_dealer = dealers[dealers["concesionario"].str.strip().str.upper() == concesion]
    periodo = _PERIODO_DASHBOARD.get(str(ws["F3"].value or "").strip().upper())
    if fila_dealer.empty or periodo is None:
        return
    codigo_dealer = int(fila_dealer["codigo_dealer"].iloc[0])
    mes_kpi = mes_referencia if periodo in ("ACUMULADO MES", "ACUMULADO MES 2S") else None

    columnas = {"BMW": "G", "MINI": "I"}
    for marca, col in columnas.items():
        k = dashboard.kpi_bloque(resumen, codigo_dealer, marca, periodo, mes_kpi, remarketing_disponible)
        ws[f"{col}12"].value = k["objetivo_retail"]
        ws[f"{col}13"].value = k["realizado_retail"]
        ws[f"{col}15"].value = k["bps"]
        ws[f"{col}17"].value = k["remarketing"] if remarketing_disponible else None
        ws[f"{col}19"].value = k["bev"]


def rellenar_plantilla(resumen: pd.DataFrame, dealers: pd.DataFrame, remarketing_disponible: bool, mes_referencia: str) -> bytes:
    """Devuelve los bytes del Excel de la plantilla original, con las
    celdas de Realizado/Objetivo rellenas con lo que la app calculó y
    el mes de referencia de cada hoja (para las columnas "Acumulado
    Mes"/"S2") sincronizado con `mes_referencia`.
    Lanza FileNotFoundError si no hay plantilla disponible."""
    if not plantilla_disponible():
        raise FileNotFoundError(f"No se encontró la plantilla en {PLANTILLA_PATH}")

    wb = openpyxl.load_workbook(PLANTILLA_PATH, data_only=False)

    objetivos = storage.read_table("objetivos")
    mercado = storage.read_table("mercado_menos_6_anos")
    mys = storage.read_table("mystery_shopping")

    for marca, hoja in _HOJAS_RETAIL_OBJ.items():
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        columnas = _localizar_columnas(ws, fila_mes=3, fila_subheader=4)
        filas = _filas_por_dealer(ws, fila_inicio=5)
        obj_marca = objetivos[objetivos["metrica"] == "Retail"] if not objetivos.empty else objetivos
        objetivo_vals = _dict_manual(obj_marca, marca, "mes")
        _escribir_mes(ws, filas, columnas, "OBJ", objetivo_vals)
        _escribir_mes(ws, filas, columnas, "RE", _dict_resumen(resumen, marca, "retail"))

    for marca, (hoja, etiqueta_bps, etiqueta_rmk) in _HOJAS_BPS_REMARKETING.items():
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        columnas = _localizar_columnas(ws, fila_mes=5, fila_subheader=6)
        filas = _filas_por_dealer(ws, fila_inicio=7)
        _escribir_mes(ws, filas, columnas, etiqueta_bps, _dict_resumen(resumen, marca, "bps"))
        # Si no hay columna "Canal Actual" en el archivo de ventas no se
        # sabe el remarketing de ningún mes -se blanquea todo (valores={})
        # en vez de dejar el dato antiguo de la plantilla o inventar un 0.
        remarketing_vals = _dict_resumen(resumen, marca, "remarketing") if remarketing_disponible else {}
        _escribir_mes(ws, filas, columnas, etiqueta_rmk, remarketing_vals)

    for marca, hoja in _HOJAS_BEV.items():
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        columnas = _localizar_columnas(ws, fila_mes=5, fila_subheader=6)
        filas = _filas_por_dealer(ws, fila_inicio=7)
        _escribir_mes(ws, filas, columnas, "BEV", _dict_resumen(resumen, marca, "bev"))

    for marca, hoja in _HOJAS_WHOLESALE.items():
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        columnas = _localizar_columnas(ws, fila_mes=5, fila_subheader=6)
        filas = _filas_por_dealer(ws, fila_inicio=7)
        _escribir_mes(ws, filas, columnas, "UC", _dict_resumen(resumen, marca, "wholesale_uc"))
        _escribir_mes(ws, filas, columnas, "YUC", _dict_resumen(resumen, marca, "wholesale_yuc"))

    for marca, hoja in _HOJAS_MERCADO.items():
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        columnas = _localizar_columnas(ws, fila_mes=5, fila_subheader=6)
        filas = _filas_por_dealer(ws, fila_inicio=7)
        mercado_vals = _dict_manual(mercado, marca, "mes")
        _escribir_mes(ws, filas, columnas, "MDO < 6 AÑOS", mercado_vals)

    if _HOJA_MYS in wb.sheetnames:
        ws = wb[_HOJA_MYS]
        filas = _filas_por_dealer(ws, fila_inicio=5)
        columnas_mys = {6: ("BMW", "S1"), 7: ("MINI", "S1"), 8: ("BMW", "S2"), 9: ("MINI", "S2")}
        for col, (marca, semestre) in columnas_mys.items():
            if not mys.empty:
                sub = mys[(mys["marca"] == marca) & (mys["semestre"] == semestre)]
                valores = {int(r["codigo_dealer"]): r["valor"] for _, r in sub.iterrows()}
            else:
                valores = {}
            for codigo, fila in filas.items():
                valor = valores.get(codigo)
                ws.cell(row=fila, column=col).value = None if valor is None else valor / 100

    _sincronizar_mes_referencia(wb, mes_referencia)
    _rellenar_dealer_dashboard(wb, resumen, dealers, remarketing_disponible, mes_referencia)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
