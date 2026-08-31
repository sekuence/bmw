"""Importa desde el Excel de seguimiento original (`SEGUIMIENTO UC
RETAIL & WHOLESALE`) todos los datos que la app no puede calcular sola
desde el archivo de ventas -objetivos, tamaño de mercado <6 años y
mystery shopping-, para no tener que volver a teclearlos a mano
concesionario por concesionario. Sólo lee valores -no fórmulas ni
vínculos externos-, así que no importa que ese archivo enlace a otros
ficheros que no tenemos."""
import openpyxl
import pandas as pd

from . import config

# (hoja, columna de inicio del primer mes, ancho de cada bloque mensual).
# El primer valor de cada bloque mensual es siempre el que nos interesa
# (OBJ/Objetivo en las hojas de objetivos, "Mdo < 6 años" en las de
# penetración de mercado). Los anchos de bloque están sacados de las
# fórmulas de la pestaña "Dealer Dashboard" del propio Excel original.
_HOJAS_OBJETIVO = {
    ("BMW", "Retail"): ("UC BMW 2026 BPS", 6, 6),
    ("MINI", "Retail"): ("UC MINI 2026 MINI NEXT", 6, 7),
    ("BMW", "BEV"): ("BEV BMW 2026", 6, 3),
    ("MINI", "BEV"): ("BEV MINI 2026", 6, 3),
}

# Hojas "PENETRACION MERCADO VO BMW/MINI": mismo formato de bloques
# mensuales que las de objetivos, primer valor del bloque = "Mdo < 6 años".
_HOJAS_MERCADO = {
    "BMW": ("PENETRACION MERCADO VO BMW", 6, 3),
    "MINI": ("PENETRACION MCDO VO MINI", 6, 3),
}

# Hoja "MYS 2026": una única pestaña con las dos marcas y los dos
# semestres en columnas fijas (no por bloques mensuales).
_HOJA_MYS = "MYS 2026"
_COLUMNAS_MYS = {  # columna (1-index) -> (marca, semestre)
    6: ("BMW", "S1"),
    7: ("MINI", "S1"),
    8: ("BMW", "S2"),
    9: ("MINI", "S2"),
}


class SeguimientoFileError(ValueError):
    pass


def _leer_bloques_mensuales(wb, hoja: str, col_inicio: int, ancho_bloque: int) -> list[tuple[int, str, float]]:
    """Devuelve [(codigo_dealer, mes, valor), ...] leyendo el primer
    valor de cada bloque mensual de `ancho_bloque` columnas, empezando
    en `col_inicio`. Las filas de resumen ("TOTAL DISTRITO X"...) se
    descartan solas porque su "código" no es un entero."""
    if hoja not in wb.sheetnames:
        return []
    ws = wb[hoja]
    filas = []
    for fila in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=col_inicio + 12 * ancho_bloque):
        codigo_dealer = fila[1].value  # columna B
        if not isinstance(codigo_dealer, int):
            continue
        for i, mes in enumerate(config.MESES):
            col = col_inicio + i * ancho_bloque
            if col > len(fila):
                break
            valor = fila[col - 1].value
            if isinstance(valor, (int, float)):
                filas.append((codigo_dealer, mes, float(valor)))
    return filas


def importar_objetivos(file_obj) -> pd.DataFrame:
    """Devuelve un DataFrame largo (codigo_dealer, marca, metrica, mes,
    valor) con todos los objetivos encontrados."""
    wb = _abrir(file_obj)
    partes = []
    for (marca, metrica), (hoja, col_inicio, ancho_bloque) in _HOJAS_OBJETIVO.items():
        datos = _leer_bloques_mensuales(wb, hoja, col_inicio, ancho_bloque)
        if datos:
            partes.append(pd.DataFrame(datos, columns=["codigo_dealer", "mes", "valor"]).assign(marca=marca, metrica=metrica))
    if not partes:
        return pd.DataFrame(columns=["codigo_dealer", "marca", "metrica", "mes", "valor"])
    return pd.concat(partes, ignore_index=True)[["codigo_dealer", "marca", "metrica", "mes", "valor"]]


def importar_mercado(file_obj) -> pd.DataFrame:
    """Devuelve un DataFrame largo (codigo_dealer, marca, mes, valor)
    con el tamaño de mercado <6 años encontrado en las pestañas de
    penetración de mercado."""
    wb = _abrir(file_obj)
    partes = []
    for marca, (hoja, col_inicio, ancho_bloque) in _HOJAS_MERCADO.items():
        datos = _leer_bloques_mensuales(wb, hoja, col_inicio, ancho_bloque)
        if datos:
            partes.append(pd.DataFrame(datos, columns=["codigo_dealer", "mes", "valor"]).assign(marca=marca))
    if not partes:
        return pd.DataFrame(columns=["codigo_dealer", "marca", "mes", "valor"])
    return pd.concat(partes, ignore_index=True)[["codigo_dealer", "marca", "mes", "valor"]]


def importar_mystery_shopping(file_obj) -> pd.DataFrame:
    """Devuelve un DataFrame largo (codigo_dealer, marca, semestre,
    valor) con las puntuaciones de mystery shopping. En el Excel
    original vienen como fracción 0-1 (0.95 = 95%); aquí se guardan
    como puntuación 0-100, igual que la entrada manual de la app."""
    wb = _abrir(file_obj)
    if _HOJA_MYS not in wb.sheetnames:
        return pd.DataFrame(columns=["codigo_dealer", "marca", "semestre", "valor"])
    ws = wb[_HOJA_MYS]
    filas = []
    for fila in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=max(_COLUMNAS_MYS)):
        codigo_dealer = fila[1].value  # columna B
        if not isinstance(codigo_dealer, int):
            continue
        for col, (marca, semestre) in _COLUMNAS_MYS.items():
            valor = fila[col - 1].value
            if isinstance(valor, (int, float)):
                filas.append({"codigo_dealer": codigo_dealer, "marca": marca, "semestre": semestre, "valor": float(valor) * 100})
    return pd.DataFrame(filas, columns=["codigo_dealer", "marca", "semestre", "valor"])


def _abrir(file_obj):
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    try:
        return openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise SeguimientoFileError(f"No se pudo abrir el archivo: {exc}") from exc


def importar_todo(file_obj) -> dict[str, pd.DataFrame]:
    """Lee de un tirón todo lo importable del Excel de seguimiento
    original: objetivos, mercado <6 años y mystery shopping. Cada
    DataFrame puede salir vacío si no se encontró la pestaña
    correspondiente -no es un error, cada dato es independiente."""
    resultado = {
        "objetivos": importar_objetivos(file_obj),
        "mercado": importar_mercado(file_obj),
        "mystery_shopping": importar_mystery_shopping(file_obj),
    }
    if all(df.empty for df in resultado.values()):
        raise SeguimientoFileError(
            "No se encontró ninguna pestaña reconocida en este archivo "
            "(objetivos, penetración de mercado o mystery shopping)."
        )
    return resultado
