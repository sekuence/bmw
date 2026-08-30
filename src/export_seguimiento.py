"""Genera el Excel con TODO el 'Seguimiento UC Retail & Wholesale'
-una hoja por pestaña x marca, tal cual se ve en la página-, no sólo
un resumen suelto (eso ya lo hace export.py). Estilo con los colores de
marca (azul BMW / naranja MINI, igual que en la app) y bordes/bandas
para que se lea como una tabla real, no como un volcado de datos."""
import io
import itertools

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import theme

FUENTE = "Arial"
COLOR_NEUTRO = "1F3864"
HEADER_FONT = Font(name=FUENTE, bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
SUBHEADER_FONT = Font(name=FUENTE, bold=True, size=10)
BASE_FONT = Font(name=FUENTE, size=10)
BASE_FONT_DESTACADO = Font(name=FUENTE, size=10, bold=True)
FILA_PAR_FILL = PatternFill("solid", fgColor="F2F2F2")
FILA_DIRECTO_FILL = PatternFill("solid", fgColor="FFF3CD")
BORDE_GRIS = Side(style="thin", color="BFBFBF")
BORDE = Border(left=BORDE_GRIS, right=BORDE_GRIS, top=BORDE_GRIS, bottom=BORDE_GRIS)
FORMATO_PORCENTAJE = '0.0"%"'  # el valor ya viene multiplicado x100 (92.3, no 0.923)

_CARACTERES_INVALIDOS = str.maketrans("", "", "[]:*?/\\")


def _nombre_hoja(titulo: str) -> str:
    return titulo.translate(_CARACTERES_INVALIDOS)[:31]


def _color_de(marca: str | None) -> str:
    if marca == "BMW":
        return theme.COLOR_BMW.lstrip("#")
    if marca == "MINI":
        return theme.COLOR_MINI.lstrip("#")
    return COLOR_NEUTRO


def _escribir_tabla(ws, df: pd.DataFrame, marca: str | None = None) -> None:
    """Escribe un DataFrame con columnas MultiIndex (nivel0=periodo,
    nivel1=submétrica) con cabecera de 2 filas, fusionando las celdas
    repetidas del nivel0 -igual que los bloques de mes del Excel
    original-, coloreada con el color de la marca."""
    header_fill = PatternFill("solid", fgColor=_color_de(marca))
    columna_id = df.columns[1][1] if len(df.columns) > 1 else ""  # "Concesionario" / "Grupo propietario" / "Marca"

    col = 1
    for nivel0, grupo in itertools.groupby(df.columns, key=lambda c: c[0]):
        span = len(list(grupo))
        c = ws.cell(row=1, column=col, value=nivel0 or None)
        c.font = HEADER_FONT
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDE
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        else:
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += span

    columnas_pct = set()
    for j, (_, nivel1) in enumerate(df.columns, start=1):
        c = ws.cell(row=2, column=j, value=nivel1)
        c.font = SUBHEADER_FONT
        c.fill = SUBHEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDE
        if str(nivel1).startswith("%"):
            columnas_pct.add(j)

    for i, (_, fila) in enumerate(df.iterrows(), start=3):
        es_directo = str(fila.iloc[1]) == "BMW DIRECTO" if len(fila) > 1 else False
        fill = FILA_DIRECTO_FILL if es_directo else (FILA_PAR_FILL if i % 2 == 0 else None)
        font = BASE_FONT_DESTACADO if es_directo else BASE_FONT
        for j, valor in enumerate(fila, start=1):
            valor = None if pd.isna(valor) else valor
            c = ws.cell(row=i, column=j, value=valor)
            c.font = font
            c.border = BORDE
            if fill:
                c.fill = fill
            if j in columnas_pct and isinstance(valor, (int, float)):
                c.number_format = FORMATO_PORCENTAJE
            if j > 2:
                c.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "C3"
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = max(14, min(28, len(str(columna_id)) + 12))
    for j in range(3, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 9
    ws.sheet_view.showGridLines = False


def build_workbook(tablas: dict[tuple[str, str], pd.DataFrame], titulos: dict[str, callable], dealers: pd.DataFrame) -> bytes:
    """tablas: {(marca, tipo): DataFrame} -una por pestaña de la app-.
    titulos: {tipo: función(marca) -> título de la pestaña}."""
    wb = Workbook()
    wb.remove(wb.active)

    nombres_usados = set()
    for (marca, tipo), df in tablas.items():
        nombre = _nombre_hoja(titulos[tipo](marca))
        while nombre in nombres_usados:
            nombre = f"{nombre[:29]}·"
        nombres_usados.add(nombre)
        ws = wb.create_sheet(nombre)
        _escribir_tabla(ws, df, marca)

    ws_maestro = wb.create_sheet("MAESTRO")
    header_fill = PatternFill("solid", fgColor=COLOR_NEUTRO)
    headers = list(dealers.columns)
    for j, h in enumerate(headers, start=1):
        c = ws_maestro.cell(row=1, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = header_fill
        c.border = BORDE
        c.alignment = Alignment(horizontal="center")
    for i, (_, fila) in enumerate(dealers.iterrows(), start=2):
        for j, valor in enumerate(fila, start=1):
            valor = None if pd.isna(valor) else valor
            c = ws_maestro.cell(row=i, column=j, value=valor)
            c.font = BASE_FONT
            c.border = BORDE
            if i % 2 == 0:
                c.fill = FILA_PAR_FILL
    ws_maestro.freeze_panes = "A2"
    for j in range(1, len(headers) + 1):
        ws_maestro.column_dimensions[get_column_letter(j)].width = 16
    ws_maestro.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
